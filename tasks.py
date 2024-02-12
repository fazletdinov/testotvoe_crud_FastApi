import asyncio
import hashlib
import json
from datetime import timedelta
from pathlib import Path
from typing import Any
from uuid import uuid4

import pandas as pd
from celery import Celery
from openpyxl import load_workbook
from sqlalchemy import String, create_engine
from sqlalchemy.dialects.postgresql import UUID

from src.core.config import settings
from src.database.redis_cache import RedisDB, get_redis

engine = create_engine(f'postgresql://{settings.db.user}:{settings.db.password.get_secret_value()}@'
                       f'{settings.db.host}:{settings.db.port}/{settings.db.name}')
celery = Celery('tasks',
                broker=settings.rabbitmq.broker_url,
                backend=settings.redis.backend_url)
celery.conf.beat_schedule = {
    'update_database': {
        'task': 'tasks.update_database',
        'schedule': timedelta(seconds=15),
    },
}

ADMIN_FILE_MENU = Path('src/admin/Menu.xlsx')
HASH_FILE_MENU = Path('src/admin/hash')


def generating_a_hash_amount() -> str:
    with ADMIN_FILE_MENU.open('rb') as file:
        hsh = hashlib.sha256()
        while True:
            data = file.read()
            if not data:
                break
            hsh.update(data)
    return hsh.hexdigest()


def read_hash() -> str:
    with HASH_FILE_MENU.open('r') as f:
        return f.read()


def write_hash(hash_summ: str) -> None:
    with HASH_FILE_MENU.open('w') as f:
        f.write(hash_summ)


def excel_to_json(array: list[list]) -> tuple[dict[str, dict[Any, Any]], dict[str, dict[Any, Any]], dict[str, dict[Any, Any]]]:
    menu_json: dict[str, dict] = {'id': {}, 'title': {}, 'description': {}}
    submenu_json: dict[str, dict] = {'id': {}, 'menu_id': {}, 'title': {}, 'description': {}}
    dish_json: dict[str, dict] = {'id': {}, 'submenu_id': {},
                                  'title': {}, 'description': {}, 'price': {}, 'discount': {}}

    menu_count, sub_count, dish_count = 0, 0, 0

    current_menu_id = ''
    current_sub_id = ''

    for row in array:
        if bool(row[0]) and bool(row[1]):
            current_menu_id = str(uuid4())
            menu_json['id'][menu_count] = current_menu_id
            menu_json['title'][menu_count] = row[1]
            menu_json['description'][menu_count] = row[2]
            menu_count += 1

        elif bool(row[0]) is False and bool(row[1]):
            current_sub_id = str(uuid4())
            submenu_json['id'][sub_count] = current_sub_id
            submenu_json['menu_id'][sub_count] = current_menu_id
            submenu_json['title'][sub_count] = row[2]
            submenu_json['description'][sub_count] = row[3]
            sub_count += 1

        elif bool(row[0]) is False and bool(row[1]) is False:
            dish_json['id'][dish_count] = str(uuid4())
            dish_json['submenu_id'][dish_count] = current_sub_id
            dish_json['title'][dish_count] = row[3]
            dish_json['description'][dish_count] = row[4]
            dish_json['price'][dish_count] = row[5]
            dish_count += 1
    return menu_json, submenu_json, dish_json


def run_update_database(data: list) -> None:
    menus_data, submenus_data, dishes_data = excel_to_json(data)

    dish_df = pd.read_json(json.dumps(dishes_data))
    dish_df.to_sql('dish', engine, if_exists='replace', index=False,
                   dtype={'id': UUID(as_uuid=True),
                          'submenu_id': UUID(as_uuid=True),
                          'title': String,
                          'description': String,
                          'price': String,
                          },
                   )

    submenu_df = pd.read_json(json.dumps(submenus_data))
    submenu_df.to_sql('submenu', engine, if_exists='replace', index=False,
                      dtype={'id': UUID(as_uuid=True),
                             'menu_id': UUID(as_uuid=True),
                             'title': String,
                             'description': String,
                             },
                      )

    menu_df = pd.read_json(json.dumps(menus_data))
    menu_df.to_sql('menu', engine, if_exists='replace', index=False,
                   dtype={'id': UUID(as_uuid=True),
                          'title': String,
                          'description': String,
                          },
                   )


@celery.task
def update_database() -> None:
    if Path(ADMIN_FILE_MENU).exists():
        new_hash = generating_a_hash_amount()
        if not Path.exists(HASH_FILE_MENU):
            write_hash('default')
        old_hash = read_hash()
        wb = load_workbook(ADMIN_FILE_MENU)
        sheet = wb.active
        data = sheet.iter_rows(values_only=True)
        if old_hash != new_hash:
            run_update_database(data)
            write_hash(new_hash)
            loop = asyncio.get_event_loop()
            redis: RedisDB = get_redis()
            if not loop.is_running():
                loop.run_until_complete(redis.delete_all())
                loop.close()
            else:
                asyncio.create_task(redis.delete_all())


# if __name__ == "__main__":
#     wb = load_workbook(ADMIN_FILE_MENU)
#     sheet = wb.active
#     data = sheet.iter_rows(values_only=True)
#     run_update_database(data)
